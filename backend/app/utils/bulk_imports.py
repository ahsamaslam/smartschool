import csv
import io
import re
from datetime import date, datetime
from typing import Dict, Iterable, List, Optional, Tuple

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook

TEACHER_IMPORT_HEADERS = [
    "school_name",
    "full_name",
    "email",
    "employee_id",
    "branch_name",
    "designation",
    "date_of_joining",
    "employment_status",
    "qualifications",
    "experience_years",
    "contact",
    "emergency_contact",
    "languages",
    "salary",
]

STUDENT_IMPORT_HEADERS = [
    "school_name",
    "full_name",
    "email",
    "student_roll_no",
    "branch_name",
    "class_name",
    "academic_session",
    "section",
    "date_of_birth",
    "gender",
    "address",
    "guardian_name",
    "primary_contact",
    "emergency_contact",
    "blood_group",
    "medical_notes",
]


def _normalize_header(value: Optional[str]) -> str:
    value = str(value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return re.sub(r"_+", "_", value).strip("_")


def _normalize_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return str(value).strip()


def _as_row_dict(headers: List[str], row_values: Iterable) -> Dict[str, object]:
    row_values = list(row_values)
    row_dict: Dict[str, object] = {}
    for idx, header in enumerate(headers):
        value = row_values[idx] if idx < len(row_values) else ""
        row_dict[header] = _normalize_cell(value)
    return row_dict


async def parse_upload_rows(file: UploadFile) -> Tuple[List[Dict[str, object]], List[str]]:
    filename = (file.filename or "").lower()
    payload = await file.read()

    if not payload:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    if filename.endswith(".csv"):
        text = payload.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        headers = [_normalize_header(h) for h in (reader.fieldnames or [])]
        if not headers:
            raise HTTPException(status_code=400, detail="Missing header row in CSV")
        rows = []
        for row in reader:
            normalized = {}
            for k, v in row.items():
                normalized[_normalize_header(k)] = _normalize_cell(v)
            rows.append(normalized)
        return rows, headers

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
        worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
        raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not raw_headers:
            raise HTTPException(status_code=400, detail="Missing header row in worksheet")
        headers = [_normalize_header(h) for h in raw_headers]
        rows = []
        for row_values in worksheet.iter_rows(min_row=2, values_only=True):
            if row_values is None:
                continue
            if not any(v not in (None, "") for v in row_values):
                continue
            rows.append(_as_row_dict(headers, row_values))
        return rows, headers

    raise HTTPException(status_code=400, detail="Unsupported file type. Please upload CSV or Excel (.xlsx/.xls)")


def build_template_bytes(
    data_headers: List[str],
    allowed_values: Dict[str, List[str]],
    sample_rows: Optional[List[List[str]]] = None,
) -> bytes:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.append(data_headers)

    for sample in sample_rows or []:
        data_sheet.append(sample)

    allowed_sheet = workbook.create_sheet(title="Allowed Values")
    allowed_sheet.append(["field", "value"])
    for field, values in allowed_values.items():
        if not values:
            allowed_sheet.append([field, ""])
            continue
        for value in values:
            allowed_sheet.append([field, value])

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalize_phone(value: object) -> str:
    return re.sub(r"\s+", "", normalize_text(value))
